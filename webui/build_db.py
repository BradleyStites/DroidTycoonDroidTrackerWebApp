#!/usr/bin/env python3
"""
Build the Droid Tycoon SQLite database from the source xlsx workbooks.

This DB is the single source of truth for the new web UI. It is a *real* SQLite
file a human can open/edit with any DB editor (DB Browser for SQLite, etc.) and
the running web server reads it live (no caching), so edits made outside the GUI
are reflected on the next request.

Inputs (exact on-disk paths; note the "Driod" typo in the filename):
  - Droid Tycoon Rebirth Tracking System/Driod Tycoon Rebirth.xlsx
      -> master table DroidRebirthDB (Super Rebirth, Rebirth, Droid Name, Droid Color)
      -> reference lists: SuperRebirthCycles (1-4), Rebirths (1-27), Droids, Colors
  - Droid Tycoon Stat Tracking/Droid Tycoon Stats.xlsx
      -> RebirthRequirementsRewards (Rebirth, Credits, Nova) cost curve

Output:
  - droid_tycoon.db  (alongside this script, in webui/)
"""

import os
import sqlite3
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)  # DroidTycoon/

REBIRTH_XLSX = os.path.join(
    ROOT, "Droid Tycoon Rebirth Tracking System", "Driod Tycoon Rebirth.xlsx"
)
STATS_XLSX = os.path.join(
    ROOT, "Droid Tycoon Stat Tracking", "Droid Tycoon Stats.xlsx"
)
DB_PATH = os.path.join(BASE, "droid_tycoon.db")

# Data-quality quarantine: these are known errors in the source sheet.
COLOR_ERROR_SUFFIX = "(Incorrect)"


def load_rows(path, sheet, max_col, header=True):
    """Read a sheet to a list of rows.

    header=True  -> skip row 1 (e.g. DroidRebirthDB, RebirthRequirementsRewards).
    header=False -> no header; data starts at row 1 (e.g. the single-column
                    reference lists SuperRebirthCycles, Rebirths, Droids, Colors).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    start = 1 if not header else 2
    out = []
    for r in range(start, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in row):
            continue
        out.append(row)
    return out


def main():
    if not os.path.exists(REBIRTH_XLSX):
        raise SystemExit(f"Missing rebirth workbook: {REBIRTH_XLSX}")
    if not os.path.exists(STATS_XLSX):
        raise SystemExit(f"Missing stats workbook: {STATS_XLSX}")

    # --- master table: DroidRebirthDB ---
    master = load_rows(REBIRTH_XLSX, "DroidRebirthDB", 5)

    droid_rebirths = []  # (super_rebirth, rebirth, droid_name, droid_color)
    dropped = 0
    for sr, rb, name, color, _ in master:
        if sr is None or rb is None or name is None:
            continue
        sr = int(sr)
        rb = int(rb)
        name = str(name).strip()
        color = str(color).strip() if color is not None else ""
        if color.endswith(COLOR_ERROR_SUFFIX) or name.endswith(COLOR_ERROR_SUFFIX):
            dropped += 1
            continue
        droid_rebirths.append((sr, rb, name, color))

    # --- upgrade chips (per rebirth tier) ---
    # Sourced from the Stats workbook's `Upgrade Chips Calculations` sheet, which
    # logs the player's accumulated Upgrade Chips at each (Super Rebirth, Rebirth)
    # stage. The *latest* snapshot for each stage is its chip total. This is the
    # per-rebirth upgrade-chip target the admin panel edits.
    chip_rows = load_rows(STATS_XLSX, "Upgrade Chips Calculations", 7, header=True)
    chips_by_stage = {}  # (sr, rb) -> chips
    for sr, rb, chips, _time, _delta, _tdelta, _permin in chip_rows:
        if sr is None or rb is None or chips is None:
            continue
        key = (int(sr), int(rb))
        val = float(chips)
        # Keep the highest (most recent / accumulated) reading for the stage.
        if key not in chips_by_stage or val > chips_by_stage[key]:
            chips_by_stage[key] = val
    upgrade_chips = [(sr, rb, chips) for (sr, rb), chips in sorted(chips_by_stage.items())]

    # --- reference lists ---
    super_cycles = sorted(
        {int(v) for v in [r[0] for r in load_rows(REBIRTH_XLSX, "SuperRebirthCycles", 1, header=False)] if v is not None}
    )
    rebirth_levels = sorted(
        {int(v) for v in [r[0] for r in load_rows(REBIRTH_XLSX, "Rebirths", 1, header=False)] if v is not None}
    )
    droid_names = [str(r[0]).strip() for r in load_rows(REBIRTH_XLSX, "Droids", 1, header=False) if r[0]]
    colors = [str(r[0]).strip() for r in load_rows(REBIRTH_XLSX, "Colors", 1, header=False) if r[0]]

    # --- cost curve: RebirthRequirementsRewards (in stats workbook) ---
    cost_rows = load_rows(STATS_XLSX, "RebirthRequirementsRewards", 3)
    costs = []  # (rebirth, credits, nova)
    for rb, credits, nova in cost_rows:
        if rb is None:
            continue
        costs.append((int(rb), credits, nova))

    # --- write DB ---
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    cur.execute(
        """
        CREATE TABLE droid_rebirths (
            id            INTEGER PRIMARY KEY,
            super_rebirth INTEGER NOT NULL,
            rebirth       INTEGER NOT NULL,
            droid_name    TEXT    NOT NULL,
            droid_color   TEXT    NOT NULL,
            UNIQUE (super_rebirth, rebirth, droid_name, droid_color)
        )
        """
    )
    cur.executemany(
        "INSERT INTO droid_rebirths (super_rebirth, rebirth, droid_name, droid_color) VALUES (?,?,?,?)",
        droid_rebirths,
    )

    cur.execute("CREATE TABLE super_rebirth_cycles (value INTEGER PRIMARY KEY)")
    cur.executemany("INSERT INTO super_rebirth_cycles (value) VALUES (?)", [(v,) for v in super_cycles])

    cur.execute("CREATE TABLE rebirth_levels (value INTEGER PRIMARY KEY)")
    cur.executemany("INSERT INTO rebirth_levels (value) VALUES (?)", [(v,) for v in rebirth_levels])

    cur.execute("CREATE TABLE droids (name TEXT PRIMARY KEY)")
    cur.executemany("INSERT INTO droids (name) VALUES (?)", [(n,) for n in droid_names])

    cur.execute("CREATE TABLE colors (name TEXT PRIMARY KEY)")
    cur.executemany("INSERT INTO colors (name) VALUES (?)", [(c,) for c in colors])

    cur.execute(
        """
        CREATE TABLE rebirth_cost (
            rebirth INTEGER PRIMARY KEY,
            credits REAL,
            nova    REAL
        )
        """
    )
    cur.executemany(
        "INSERT OR REPLACE INTO rebirth_cost (rebirth, credits, nova) VALUES (?,?,?)", costs
    )

    # Player's current stage in the grind (mutable state; the Rebirth button
    # advances this). Single row, id=1. Seeded to the starting stage SR1/R1.
    cur.execute(
        """
        CREATE TABLE player_stage (
            id            INTEGER PRIMARY KEY CHECK (id = 1),
            super_rebirth INTEGER NOT NULL,
            rebirth       INTEGER NOT NULL
        )
        """
    )
    cur.execute(
        "INSERT OR IGNORE INTO player_stage (id, super_rebirth, rebirth) VALUES (1, 1, 1)"
    )

    cur.execute(
        """
        CREATE TABLE rebirth_upgrade_chips (
            super_rebirth INTEGER NOT NULL,
            rebirth       INTEGER NOT NULL,
            chips         REAL    NOT NULL,
            PRIMARY KEY (super_rebirth, rebirth)
        )
        """
    )
    cur.executemany(
        "INSERT OR REPLACE INTO rebirth_upgrade_chips (super_rebirth, rebirth, chips) VALUES (?,?,?)",
        upgrade_chips,
    )

    con.commit()
    n_master = cur.execute("SELECT COUNT(*) FROM droid_rebirths").fetchone()[0]
    n_stages = cur.execute(
        "SELECT COUNT(DISTINCT super_rebirth || '-' || rebirth) FROM droid_rebirths"
    ).fetchone()[0]
    con.close()

    print(f"Built {DB_PATH}")
    print(f"  droid_rebirths rows : {n_master}")
    print(f"  distinct stages     : {n_stages}")
    print(f"  super cycles        : {super_cycles}")
    print(f"  rebirth levels      : {rebirth_levels[0]}..{rebirth_levels[-1]} ({len(rebirth_levels)})")
    print(f"  droids list         : {len(droid_names)}")
    print(f"  colors list         : {colors}")
    print(f"  cost rows           : {len(costs)}")
    print(f"  upgrade chip rows   : {len(upgrade_chips)}")
    print(f"  quarantined rows    : {dropped}")


if __name__ == "__main__":
    main()
