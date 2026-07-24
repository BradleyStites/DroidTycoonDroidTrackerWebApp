#!/usr/bin/env python
"""
build_data.py
Extracts the Droid Tycoon rebirth data model from the two project workbooks and
writes a normalized JSON file consumed by needed_droids_panel.html.

Source workbooks (referenced by their exact on-disk names):
  - Droid Tycoon Rebirth Tracking System/Driod Tycoon Rebirth.xlsx
        (note the filename typo "Driod" -- matches the actual file on disk)
  - Droid Tycoon Stat Tracking/Droid Tycoon Stats.xlsx

Output:
  data/rebirth_data.json

The output schema mirrors REBIRTH_DATA_SCHEMA.md:
  RebirthStage { super_rebirth, rebirth, droid_name, droid_color }
  3 rows share each (super_rebirth, rebirth) key -> the "needed droids" for a stage.
"""

import json
import os
from collections import defaultdict, OrderedDict

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
CONTENT_XLSX = os.path.join(
    ROOT, "Droid Tycoon Rebirth Tracking System", "Driod Tycoon Rebirth.xlsx"
)
STATS_XLSX = os.path.join(ROOT, "Droid Tycoon Stat Tracking", "Droid Tycoon Stats.xlsx")

# Canonical color palette from the Colors sheet (ground truth = DroidRebirthDB).
# Rows whose color ends with "(Incorrect)" are data errors -> quarantined.
INCORRECT_SUFFIX = "(Incorrect)"

VALID_COLORS = ["Base", "Default", "Gold", "Diamond", "Rainbow", "Beskar"]


def _nonempty_rows(ws):
    for row in ws.iter_rows(values_only=True):
        if row and any(c is not None for c in row):
            yield row


def _first_col(ws):
    out = []
    for row in _nonempty_rows(ws):
        out.append(row[0])
    return out


def extract_content():
    wb = openpyxl.load_workbook(CONTENT_XLSX, data_only=True, read_only=True)

    # --- DroidRebirthDB : master required-droids table ---
    db = wb["DroidRebirthDB"]
    header = None
    stages = defaultdict(list)
    quarantined = []
    for row in _nonempty_rows(db):
        if header is None:
            header = row
            continue
        sr, rb, name, color, *_ = row
        if sr is None or rb is None or name is None:
            continue
        sr = int(sr)
        rb = int(rb)
        name = str(name).strip()
        color = str(color).strip() if color is not None else ""
        if color.endswith(INCORRECT_SUFFIX):
            quarantined.append({"super_rebirth": sr, "rebirth": rb, "droid_name": name, "droid_color": color})
            continue
        stages[(sr, rb)].append({"droid_name": name, "droid_color": color})

    # Order stages by (sr, rb)
    ordered_stages = OrderedDict()
    for key in sorted(stages.keys()):
        sr, rb = key
        ordered_stages[f"{sr}-{rb}"] = {
            "super_rebirth": sr,
            "rebirth": rb,
            "droids": stages[key],
        }

    droids_list = [str(x) for x in _first_col(wb["Droids"]) if x is not None]
    colors_list = [str(x) for x in _first_col(wb["Colors"]) if x is not None]
    ranks_list = [str(x) for x in _first_col(wb["Ranks"]) if x is not None]
    super_cycles = [int(x) for x in _first_col(wb["SuperRebirthCycles"]) if x is not None]
    rebirths_domain = [int(x) for x in _first_col(wb["Rebirths"]) if x is not None]

    wb.close()

    return {
        "stages": ordered_stages,
        "droids_list": droids_list,
        "colors_list": colors_list,
        "ranks_list": ranks_list,
        "super_cycles": super_cycles,
        "rebirths_domain": rebirths_domain,
        "quarantined": quarantined,
    }


def extract_costs():
    wb = openpyxl.load_workbook(STATS_XLSX, data_only=True, read_only=True)
    ws = wb["RebirthRequirementsRewards"]
    cost_by_rebirth = {}
    header = None
    for row in _nonempty_rows(ws):
        if header is None:
            header = row
            continue
        rb, credits, nova = row[0], row[1], (row[2] if len(row) > 2 else None)
        if rb is None:
            continue
        rb = int(rb)
        cost_by_rebirth[rb] = {
            "credits": float(credits) if credits is not None else None,
            "nova": float(nova) if nova is not None else None,
        }
    wb.close()
    return cost_by_rebirth


def main():
    if not os.path.exists(CONTENT_XLSX):
        raise SystemExit(f"Content workbook not found: {CONTENT_XLSX}")
    if not os.path.exists(STATS_XLSX):
        raise SystemExit(f"Stats workbook not found: {STATS_XLSX}")

    content = extract_content()
    costs = extract_costs()

    # Attach per-stage cost (keyed by rebirth; null if no cost entry)
    for stage in content["stages"].values():
        rb = stage["rebirth"]
        stage["cost"] = costs.get(rb)

    model = {
        "source": {
            "content_workbook": "Droid Tycoon Rebirth Tracking System/Driod Tycoon Rebirth.xlsx",
            "stats_workbook": "Droid Tycoon Stat Tracking/Droid Tycoon Stats.xlsx",
            "note": "Filename 'Driod' is a typo in the original file and is preserved as-is.",
        },
        "valid_colors": VALID_COLORS,
        "droids_per_stage": 3,
        "super_cycle_range": [min(content["super_cycles"]), max(content["super_cycles"])],
        "rebirth_range": [min(content["rebirths_domain"]), max(content["rebirths_domain"])],
        "totals": {
            "super_cycles": len(content["super_cycles"]),
            "rebirths_per_cycle": len(content["rebirths_domain"]),
            "stages": len(content["stages"]),
            "droids_total": sum(len(s["droids"]) for s in content["stages"].values()),
            "distinct_droid_names": len({d["droid_name"] for s in content["stages"].values() for d in s["droids"]}),
            "quarantined_rows": len(content["quarantined"]),
        },
        "stages": content["stages"],
        "reference": {
            "droids": content["droids_list"],
            "colors": content["colors_list"],
            "ranks": content["ranks_list"],
            "super_cycles": content["super_cycles"],
            "rebirths": content["rebirths_domain"],
        },
        "cost_curve": costs,
        "quarantined": content["quarantined"],
    }

    out_dir = os.path.join(ROOT, "data")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "rebirth_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(model, f, indent=2, ensure_ascii=False)

    print(f"Wrote {out_path}")
    print(f"  stages={model['totals']['stages']} "
          f"droids_total={model['totals']['droids_total']} "
          f"distinct={model['totals']['distinct_droid_names']} "
          f"quarantined={model['totals']['quarantined_rows']}")


if __name__ == "__main__":
    main()
